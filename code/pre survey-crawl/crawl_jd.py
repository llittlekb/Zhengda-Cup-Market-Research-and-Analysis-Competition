import time
import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

search=input('请输入想要搜索的商品：').strip()
browser = webdriver.Chrome()
browser.implicitly_wait(5)
browser.get('https://search.jd.com/Search?keyword='+search)

wb=openpyxl.Workbook()
ws=wb.active
ws.title='京东商品信息' 
ws['A1']='商品标题' 
ws['B1']='商品价格' 
ws['C1']='评论数量'
ws['D1']='标签' 
ws['E1']='商品详情页' 
i=0

for page in range(0,5):
    js = 'var q=document.documentElement.scrollTop=8000' 
    browser.execute_script(js)
    # browser.implicitly_wait(20)
    time.sleep(10)
    lis = browser.find_elements(By.CSS_SELECTOR,'#J_goodsList ul li')

    for li in lis:
        title=li.find_element(By.CSS_SELECTOR,'.p-name em').text.split('\n')[-1]
        price=li.find_element(By.CSS_SELECTOR,'.p-price strong i').text
        commit=li.find_element(By.CSS_SELECTOR,'.p-commit strong a').text
        href=li.find_element(By.CSS_SELECTOR,'.p-img a').get_attribute('href')
        icons=li.find_elements(By.CSS_SELECTOR,'.p-icons i')
        icon=' '.join([i.text for i in icons])
        list=[title,price,commit,icon,href]
        print(list)
        ws.cell(i+2,1,list[0])
        ws.cell(i+2,2,list[1])
        ws.cell(i+2,3,list[2])
        ws.cell(i+2,4,list[3])
        ws.cell(i+2,5).hyperlink=list[4]
        i+=1
        
    next_button = WebDriverWait(browser, 10).until(EC.visibility_of_element_located((By.CSS_SELECTOR, '.pn-next')))
    next_button.click()

ws.column_dimensions['A'].width=80
ws.column_dimensions['B'].width=10
ws.column_dimensions['C'].width=10
ws.column_dimensions['D'].width=30
ws.column_dimensions['E'].width=50
wb.save('京东数据抓取.xlsx')
browser.quit()
